import random

import openpyxl
from rest_framework import viewsets
from rest_framework.response import Response

from api.models import Bills, Client, Organization
from api.serializers import CheckSerializer, ClientSerializer, UploadSerializer

SERVICE_CHOICE = {
    1: "консультация",
    2: "лечение",
    3: "стационар",
    4: "диагностика",
    5: "лаборатория"
}

ALLOWED_EXTENSIONS = {"xlsm", "xlsx", "xlsb", "xls"}


def allowed_file(filename: str) -> bool:
    """ Функция проверки допустимого расширения файла """
    return "." in filename and \
        filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def fraud_score(service: str) -> float:
    """Детектор мошенничества"""
    return round(random.uniform(0, 1), 2)


def address_normal(addr: str) -> str:
    """Нормализация адресса"""
    if addr:
        normaddr = 'Адрес: ' + addr
    return normaddr


def classifier_of_services(serv: str):
    """Классификатор услуг"""
    DICT_SERVICE = {
        1: "консультация",
        2: "лечение",
        3: "стационар",
        4: "диагностика",
        5: "лаборатория"
        }
    service_num = random.randint(1, 5)
    service_class = service_num,
    service_name = DICT_SERVICE[service_num]
    return service_class, service_name


def excel_data_uploade(worksheet) -> list:
    """Формирование списка данных из полученного листа Excel"""
    excel_data = list()
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        excel_data.append(row_data)
    return excel_data


def client_org(file):
    """Запись в базу моделей Client и Organization"""
    worksheet = file["client"]
    excel_data = excel_data_uploade(worksheet)
    for data in excel_data[1:]:
        if not Client.objects.filter(name=data[0]).exists():
            Client.objects.create(name=data[0])

    worksheet = file["organization"]
    excel_data = excel_data_uploade(worksheet)
    try:
        for data in excel_data[1:]:
            client_name, name, address = data
            if not Organization.objects.filter(
                client_name=Client.objects.filter(
                    name=client_name)[0], name=name).exists():

                Organization.objects.create(
                    client_name=Client.objects.filter(name=client_name)[0],
                    name=name,
                    address=address_normal(address),
                    fraud_weight=0)
    except IndexError as e:
        print(e)
    finally:
        print("Загрузка завершена")


def bills(file):
    sheets = file.sheetnames
    worksheet = file[sheets[0]]
    excel_data = excel_data_uploade(worksheet)
    for data in excel_data[1:]:
        client_name, client_org, check_number, check_sum, date, service, *last = data
        service_class, service_name = classifier_of_services(service)
        try:
            frag_num = fraud_score(service)
            Bills.objects.create(
                clinet_name=Client.objects.filter(name=client_name)[0],
                client_org=Organization.objects.filter(
                    client_name=Client.objects.filter(
                        name=client_name)[0], name=client_org)[0],
                check_number=int(float(check_number)),
                check_sum=float(check_sum),
                date=date,
                service=service,
                fraud_score=frag_num,
                service_class=service_class[0],
                service_name=service_name
            )
            if frag_num >= 0.9:
                org = Organization.objects.get(name=client_org)
                org.fraud_weight += 1
                org.save()
        except Exception as e:
            print(e)
        finally:
            print("Загрузка завершена")


def uploaded(file, file_name):
    dict_name = {"client_org": client_org, "bills": bills}
    func = dict_name[file_name]
    return func(file)


class ClientsViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer


class CheckViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Bills.objects.all()
    serializer_class = CheckSerializer
    filterset_fields = ("clinet_name", "client_org")


class UploadeViewSet(viewsets.ViewSet):
    serializer_class = UploadSerializer

    def list(self, request):
        return Response("GET API")

    def create(self, request):
        file_uploaded = request.FILES.get("file_uploaded")
        allowed = allowed_file(str(file_uploaded))
        if allowed:
            wb = openpyxl.load_workbook(file_uploaded)
            filename = str(file_uploaded).rsplit(".", 1)[0].lower()
            uploaded(wb, filename)

            content_type = file_uploaded.content_type
            response = f"Вы загрузили файл {content_type}"
            return Response(response)
        else:
            response = "Выберите файл Excel"
            return Response(response)
